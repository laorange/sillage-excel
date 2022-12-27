import datetime

from openpyxl import Workbook, styles, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image

from course import ApiHandler, CourseDecorator
from course.apiHandler import semester_start_date, max_week_num


class ExcelHandler:
    template = 'template.xlsx'
    start_col_names = ['B', 'N', "Z", "AL", "AX", "BJ", "BV"]
    end_col_names = ['M', 'Y', "AK", "AW", "BI", "BU", "CG"]
    start_cols = [column_index_from_string(col) for col in start_col_names]
    end_cols = [column_index_from_string(col) for col in end_col_names]
    rows = [8, 12, 19, 23, 29]
    date_row = 5
    total_col_span = 12

    def __init__(self, course_handler: CourseDecorator):
        self.ch: CourseDecorator = course_handler
        self.wb: Workbook = load_workbook(filename=self.template)
        self.ws: Worksheet = self.wb.active

        self.fill_with_courses()

    def fill_with_courses(self):
        query_date = semester_start_date
        for week in range(1, max_week_num + 1):
            new_ws: Worksheet = self.wb.copy_worksheet(self.ws)
            new_ws.title = f"Semaine {week}"
            new_ws.add_image(Image("./img/1.jpg"), 'A1')
            new_ws.add_image(Image("./img/2.png"), 'CA1')

            for col_index, (start_col, end_col) in enumerate(zip(self.start_cols, self.end_cols)):
                date_str = query_date.strftime("%Y-%m-%d")
                new_ws[f"{get_column_letter(start_col)}{self.date_row}"] = query_date.strftime("%m/%d/%Y")

                for row_index, row in enumerate(self.rows):
                    lesson_num = row_index + 1
                    courses = self.ch.filter_of_date(date_str).filter_of_lesson_num(lesson_num)
                    if len(courses.value) > 0:
                        col_span = self.total_col_span // len(courses.value)

                        for course_index, course in enumerate(courses.value):
                            col = start_col + course_index * col_span
                            cell = new_ws[f"{get_column_letter(col)}{row}"]
                            cell.value = str(course)
                            cell.fill = styles.PatternFill("solid", fgColor=course.info.bgc.strip("#")[:6])
                            cell.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

                            new_ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col + col_span - 1)}{row + 2}")

                            if course_index >= end_col - start_col or course_index == len(courses.value) - 1:
                                break
                    else:
                        new_ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row + 2}")

                query_date += datetime.timedelta(days=1)

            # for idx, col in enumerate(new_ws.columns, 1):
            #     new_ws.column_dimensions[get_column_letter(idx)].auto_size = True

    def output(self, file_path: str):
        self.wb.remove(self.ws)
        self.wb.save(filename=file_path)


def save_api_handler():
    import pickle
    with open("api_handler", "wb") as f:
        pickle.dump(ApiHandler(), f)


def get_api_handler(debug=False) -> ApiHandler:
    if debug:
        import pickle
        with open("api_handler", "rb") as f:
            return pickle.load(f)
    else:
        return ApiHandler()


if __name__ == '__main__':
    api_handler: ApiHandler = get_api_handler()
    url = input("请前往`迹云课表`复制网址：")  # "https://course.siae.top/#/course/?grade=17%E7%BA%A7"
    course_decorator = api_handler.get_filtered_courses(url)
    ExcelHandler(course_decorator).output("output.xlsx")
